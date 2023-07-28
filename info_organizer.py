from openpyxl import load_workbook
import os
import shutil
import json
import sys

'''
This scipt grab information from product excel file
Organize the product information needed for the website into a json file
Move product images from the source folder to the structured folder
'''


class product:
    def __init__(self, category, ID, color, price, decription):
        self.category = category
        self.ID = ID
        self.color = color
        self.price = price
        self.decription = decription


def load_dictionary(dict_json):
    # load category json
    with open(dict_json, 'r') as file:
        loaded_dict = json.load(file)
        return loaded_dict


def get_img_directory(category_dict, item):
    # The directory to store single product images
    category_code = category_dict[item["category"]]["img_directory_keyword"]
    directory = "-".join(['images', category_code]) + \
        '\\'+item["category"]+'\\'+item["ID"]
    # print(f"image directory {directory}")
    return '.\\'+directory


def copy_image_to_folder(image_path, folder_path):
    # copy image to the target folder
    shutil.copy(image_path, folder_path)


def get_profile_img_path(item, category_dict):
    img_directory = get_img_directory(
        category_dict, item)+'\\'+item["ID"]+'_profile.jpg'
    # print(img_directory)
    return img_directory


def copy_example_profile_img(target_img_dir):
    '''
    An item's profile page on its category page
    It may have requirement on image size or content. 
    So copy a default profile image and manually check/replace later.  
    '''
    source_img = 'profile_example.jpg'  # hardcoded
    abs_target = os.path.abspath(target_img_dir)
    shutil.copy(source_img, abs_target)
    print(f"{target_img_dir} profile img copied")


def find_images_with_word(category_dict, source_folder, itemID, items_dict):
    images_to_copy = []
    # find images with itemID in source folder
    for file_name in os.listdir(source_folder):
        if itemID in file_name:
            if file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                images_to_copy.append(file_name)
    # create target folder
    target_path = get_img_directory(category_dict, items_dict[itemID])
    if not os.path.exists(target_path):
        os.makedirs(target_path)

    # copy images to the folder
    for image_file in images_to_copy:
        image_path = os.path.join(source_folder, image_file)
        copy_image_to_folder(image_path, target_path)

    profile_img_dir = get_profile_img_path(items_dict[itemID], category_dict)
    copy_example_profile_img(profile_img_dir)


def get_worksheet(workbook):
    # if there are multiple work sheets, ask user choose one
    defult_worksheet_id = 0
    if(len(workbook.sheetnames) > 1):
        defult_worksheet_id = int(
            input("Enter the worksheet number: "))-1
    # get worksheet
    return workbook[workbook.sheetnames[defult_worksheet_id]]


def get_item_price(worksheet):
    # Case 1: All items in the same Excel file have the same category, price, and description.
    # Get item Price
    price_cell = input("Enter the price's cell: ")
    price_col = ord(price_cell[0])-ord("a")
    price_row = int(price_cell[1:])-1
    price = list(worksheet.rows)[price_row][price_col].value
    if type(price) == str:
        price = round(float(price.strip('$')))
    return price


def get_ietm_description(worksheet):
    # Get item description
    description_cell = input("Enter the description's cell: ")
    description_col = ord(description_cell[0])-ord("a")
    description_row = int(description_cell[1:])-1
    # descriptions may contain characters with unsupported encoding.
    return list(worksheet.rows)[description_row][description_col].value.encode(
        'ascii', 'ignore').decode('utf-8')


def import_excel_file(excel_file):
    # import excel file
    # load workbook
    wb = load_workbook(excel_file)
    ws = get_worksheet(wb)

    price = get_item_price(ws)
    description = get_ietm_description(ws)

    # ask user input where the data locate
    data_start = int(
        input("Enter the row number where the data starts: "))-1
    data_end = int(input("Enter the row number where the data ends: "))
    # load all selected rows
    data_rows = list(ws.rows)[data_start:data_end]
    print(f"Found {len(data_rows)} rows of data.")

    items_dict = extract_cols_info(data_rows, price, description)
    return items_dict


def extract_cols_info(data_rows, price, description):
    # extrat info from data rows
    items_dict = {}
    item_col = ord(input("Enter the item number's column: "))-ord("a")
    dec_col = ord(input("Enter the dec's column: "))-ord("a")
    color_col = ord(input("Enter the color's column:"))-ord("a")

    for row in data_rows:
        # create full dictionary entry
        items_dict.update({row[item_col].value: product(
                          row[dec_col].value.replace("  ", "_").replace(" ", "_"), row[item_col].value, row[color_col].value, price*2, description).__dict__})
    return items_dict


def organize_images(category_dict, data_dict, images_folder):
    items_dict = data_dict
    # find and move images in target folder
    for key in items_dict.keys():
        find_images_with_word(category_dict, images_folder, key, items_dict)
        print(f"Images for the item '{key}' moved to the folder.")
    print("Image files moved into respective folders successfully!")


def export_information(items_dict):
    with open('item_dict.json', 'w') as file:
        json.dump(items_dict, file)


'''
===  Main ===
- load categroy dictionary
- read product data from excel file
- move product images into corresponding folder
- export oegaznied product infomation into json file
Argument:
- excel_file: product excel file
- image_source: the folder that contains all the product images ( in my case)
- category_dict: category json file
'''


def main(excel_file, image_source_folder, category_json):
    data_file = excel_file
    images_folder = image_source_folder  # item images folder
    category_dict = load_dictionary(category_json)

    # get data in excel
    data_dict = import_excel_file(data_file)

    # process data
    organize_images(category_dict, data_dict, images_folder)
    export_information(data_dict)


'''
python info_organizer.py
Argument: excel_file, image_source, category_json
'''
main('file.xlsx', 'S://', 'category_dict.json')
