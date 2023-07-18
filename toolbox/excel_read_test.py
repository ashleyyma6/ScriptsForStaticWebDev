from openpyxl import load_workbook
import pandas as pd

'''
This script test if the content of merged cells can be read, all the cells have same value?
Result: No
Current Solution: Manually umerge and copy cell values. 
'''


def print_2d_array(array):
    for row in array:
        for item in row:
            print(item, end='\t')  # Separate items with a tab
        print()  # Move to the next line for the next row


file = ''

wb = load_workbook(file)  # load entire work book

# if there are multiple work sheets
defult_worksheet_id = 0

# ask for user input where the data locate
data_start = int(
    input("Please enter the row number where the data starts: "))-1
data_end = int(input("Please enter the row number where the data ends: "))

ws = wb[wb.sheetnames[defult_worksheet_id]]  # load 1 worksheets
data_rows = list(ws.rows)[data_start:data_end]  # load all rows
print(f"Found {len(data_rows)} rows of data.")
print_2d_array(data_rows)

print(f"Test Panda csv library print")
df = pd.DataFrame(pd.read_excel(file))
print(df)
